from __future__ import annotations

from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from .models import LiquidacionDoc, MercaderiaEntregadaItem


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel_cpns(docs: List[LiquidacionDoc]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "CPNs"

    # Hoja principal (CPNs)
    headers = [
        "Fecha",
        "COE",
        "Comprobante",      # SOLO 3302-29912534
        "Acopio/Comprador",
        "CUIT Comprador",
        "Tipo de grano",
        "Campaña",
        "Kilos",
        "Precio",           # formato $
        "Subtotal",
        "IVA",
        "Total",
    ]
    ws.append(headers)

    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for d in docs:
        row = [
            d.fecha or "",
            d.coe or "",
            d.comprobante or "",
            d.comprador_rs or "",
            (d.comprador_cuit or "").replace("-", ""),
            d.grano or "",
            d.campaña or "",
            float(d.kilos or 0),
            float(d.precio_kg or 0),
            float(d.subtotal or 0),
            float(d.iva or 0),
            float(d.total or 0),
        ]
        ws.append(row)

    # Formatos CPNs
    col_idx = {h: i + 1 for i, h in enumerate(headers)}
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=col_idx["Kilos"]).number_format = '#,##0.00'
        ws.cell(row=r, column=col_idx["Precio"]).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=col_idx["Subtotal"]).number_format = '#,##0.00'
        ws.cell(row=r, column=col_idx["IVA"]).number_format = '#,##0.00'
        ws.cell(row=r, column=col_idx["Total"]).number_format = '#,##0.00'
        ws.cell(row=r, column=col_idx["CUIT Comprador"]).number_format = "0"

    _set_col_widths(ws, [12, 14, 16, 40, 14, 18, 12, 12, 12, 14, 14, 14])

    # Hoja Mercadería Entregada
    ws2 = wb.create_sheet("Mercadería Entregada")
    headers2 = [
        "Fecha",
        "Nro Comprobante",   # repetir aquí
        "Acopio/Comprador",
        "Tipo de grano",
        "Campaña",
        "Kilos",
        "Precio",            # $
        "Observación",
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = bold
        cell.alignment = Alignment(vertical="center")
    ws2.freeze_panes = "A2"

    for d in docs:
        items: List[MercaderiaEntregadaItem] = d.mercaderia_entregada or []
        if not items:
            # si no hay sección detectada, igual podés querer un renglón base
            continue

        for it in items:
            ws2.append([
                it.fecha or d.fecha or "",
                d.comprobante or "",
                d.comprador_rs or "",
                it.grano or d.grano or "",
                it.campaña or d.campaña or "",
                float(it.kilos or 0),
                float(it.precio or d.precio_kg or 0),
                it.observacion or "",
            ])

    # formatos
    col_idx2 = {h: i + 1 for i, h in enumerate(headers2)}
    for r in range(2, ws2.max_row + 1):
        ws2.cell(row=r, column=col_idx2["Kilos"]).number_format = '#,##0.00'
        ws2.cell(row=r, column=col_idx2["Precio"]).number_format = '"$"#,##0.00'

    _set_col_widths(ws2, [12, 16, 40, 18, 12, 12, 12, 28])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
