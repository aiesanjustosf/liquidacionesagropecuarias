from __future__ import annotations
from io import BytesIO
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .parser_liquidaciones import LiquidacionDoc

def _pv_nro_from_coe(coe: Optional[str]) -> Tuple[Optional[int], Optional[int]]:
    if not coe or len(coe) < 12:
        return None, None
    return int(coe[:4]), int(coe[4:12])

def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_excel_gastos(docs: List[LiquidacionDoc]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    headers = [
        "Fecha Emisión", "Fecha Recepción", "Cpbte", "Tipo", "Suc.", "Número",
        "Razón Social/Denominación Proveedor", "Tipo Doc.", "CUIT", "Domicilio",
        "C.P.", "Pcia", "Cond Fisc", "Cód. Neto", "Neto Gravado",
        "Alic.", "IVA Liquidado", "IVA Crédito", "Cód NG/Ex", "Conceptos NG/EX"
    ]
    ws.append(headers)
    _set_widths(ws, [14,14,8,8,7,12,45,10,14,40,6,6,10,10,14,8,14,14,12,16])

    amt_fmt = "#.##0,00"
    aliq_fmt = "#.##0,000"

    for d in docs:
        pv, nro = _pv_nro_from_coe(d.coe)

        prov_rs = d.comprador_rs
        prov_cuit = d.comprador_cuit
        prov_dom = d.comprador_dom
        prov_cf = d.comprador_cf

        deducs = d.deducciones
        if not deducs:
            continue

        for it in deducs:
            alic = it.alicuota or 0.0
            tipo = 202 if abs(alic - 21.0) < 0.0001 else 203

            cpbte = "ND"  # ajuste pedido

            cod_neto = tipo
            neto = it.base
            iva = it.iva

            cod_ngex = None
            ex = it.exento
            if ex is not None and ex != 0:
                cod_ngex = 203

            ws.append([
                d.fecha, d.fecha, cpbte, tipo, pv, nro,
                prov_rs, 80, prov_cuit, prov_dom,
                None, None, prov_cf,
                cod_neto, neto,
                alic if alic != 0 else None,
                iva, iva,
                cod_ngex, ex
            ])

    for r in range(2, ws.max_row + 1):
        for c in [15,17,18,20]:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = amt_fmt
        cell = ws.cell(row=r, column=16)
        if isinstance(cell.value, (int, float)):
            cell.number_format = aliq_fmt

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
