from __future__ import annotations
from io import BytesIO
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .parser_liquidaciones import LiquidacionDoc

COD_NETO = {
    "Soja": 123,
    "Maíz": 124,
    "Trigo": 161,
    "Girasol": 157,
    "Arveja": 120,
    "Sorgo": 151,
    "Camelina Sativa": 162,
}

def _pv_nro_from_coe(coe: Optional[str]) -> Tuple[Optional[int], Optional[int]]:
    if not coe or len(coe) < 12:
        return None, None
    return int(coe[:4]), int(coe[4:12])

def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_excel_ventas(docs: List[LiquidacionDoc]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    headers = [
        "Fecha Emisión", "Fecha Recepción", "Cpbte", "Letra", "Suc.", "Número",
        "Razón Social/Denominación Cliente", "Tipo Doc.", "CUIT", "Domicilio",
        "C.P.", "Pcia", "Cond Fisc", "Cód. Neto", "Neto Gravado",
        "Alic.", "IVA Liquidado", "IVA Débito", "Cód NG/EX", "Conceptos NG/EX"
    ]
    ws.append(headers)
    _set_widths(ws, [14,14,8,6,7,12,45,10,14,40,6,6,10,10,14,8,14,14,12,16])

    amt_fmt = "#.##0,00"
    aliq_fmt = "#.##0,000"

    for d in docs:
        pv, nro = _pv_nro_from_coe(d.coe)
        cod_neto = COD_NETO.get(d.grano or "", None)

        # Línea principal: comprador únicamente
        ws.append([
            d.fecha, d.fecha, d.tipo_comprobante, "A", pv, nro,
            d.comprador_rs, 80, d.comprador_cuit, d.comprador_dom,
            None, None, d.comprador_cf, cod_neto,
            d.subtotal, d.alicuota_iva, d.iva, d.iva,
            None, None
        ])

        def add_ret(code: str, amount: float):
            ws.append([
                d.fecha, d.fecha, "RV", "A", pv, nro,
                d.comprador_rs, 80, d.comprador_cuit, d.comprador_dom,
                None, None, d.comprador_cf, cod_neto,
                amount, None, None, None,
                code, None
            ])

        # Retenciones: una línea por retención
        if d.ret_iva is not None and d.ret_iva != 0:
            add_ret("RA07", d.ret_iva)
        if d.ret_gan is not None and d.ret_gan != 0:
            add_ret("RA05", d.ret_gan)

    # Formatos
    for r in range(2, ws.max_row + 1):
        for c in [15,17,18,20]:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = amt_fmt
        cell = ws.cell(row=r, column=16)
        if isinstance(cell.value, (int, float)):
            cell.number_format = aliq_fmt

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
