from __future__ import annotations
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .parser_liquidaciones import LiquidacionDoc

def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_excel_cpns(docs: List[LiquidacionDoc]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "CPNs"

    headers = [
        "FECHA", "COMPROBANTE", "ACOPIO", "TIPO DE GRANO", "CAMPAÑA",
        "CANTIDAD DE KILOS", "PRECIO", "LOCALIDAD",
        "ME - Nro comprobante", "ME - Procedencia", "ME - Peso (kg)", "ME - Puerto",
        "ME - Grado", "ME - Factor", "ME - Contenido Proteico"
    ]
    ws.append(headers)
    _set_widths(ws, [12,18,40,16,14,18,12,18,18,25,14,20,10,10,18])

    amt_fmt = "#,##0.00"

    for d in docs:
        pv = (d.coe or "")[:4]
        nro = (d.coe or "")[4:12] if d.coe and len(d.coe) >= 12 else ""
        comprobante = f"{d.tipo_comprobante}A {pv}-{nro}".strip()

        me = d.mercaderia_entregada
        ws.append([
            d.fecha,
            comprobante,
            d.comprador_rs,
            d.grano,
            d.campania or "",
            d.kilos,
            d.precio_kg,
            d.localidad,
            me.nro_comprobante,
            me.procedencia,
            me.peso_kg,
            me.puerto,
            me.grado,
            me.factor,
            me.contenido_proteico,
        ])

    for r in range(2, ws.max_row + 1):
        # Kilos, precio, peso mercadería entregada
        for c in [6, 7, 11]:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = amt_fmt

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
