# -*- coding: utf-8 -*-
from __future__ import annotations

from io import BytesIO
from typing import List, Dict, Any, Optional

import math
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from parser import Liquidacion

# Excel format codes (internos). En Excel AR se ven como 1.000,00 y 10,000.
FMT_NUM2  = "#,##0.00"        # 1.000,00
FMT_NUM3  = "0.000"           # 10,000
FMT_MONEY = '"$"#,##0.00'     # $1.000,00
FMT_CUIT  = "0"


VENTAS_COLUMNS = [
    "Fecha dd/mm/aaaa","Cpbte","Tipo","Suc.","Número",
    "Razón Social o Denominación Cliente ",
    "Tipo Doc.","CUIT","Domicilio","C.P.","Pcia","Cond Fisc",
    "Cód. Neto","Neto Gravado","Alíc.","IVA Liquidado","IVA Débito",
    "Cód. NG/EX","Conceptos NG/EX","Cód. P/R","Perc./Ret.","Pcia P/R","Total"
]

# Sin columna "Tipo" (3ra) como pediste
COMPRAS_COLUMNS = [
    "Fecha Emisión ","Fecha Recepción","Cpbte","Suc.","Número",
    "Razón Social/Denominación Proveedor",
    "Tipo Doc.","CUIT","Domicilio","C.P.","Pcia","Cond Fisc",
    "Cód. Neto","Neto Gravado","Alíc.","IVA Liquidado","IVA Crédito",
    "Cód. NG/EX","Conceptos NG/EX","Cód. P/R","Perc./Ret.","Pcia P/R","Total"
]


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _nan_to_none(v):
    if v is None:
        return None
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return None
    return v


def _append_df(ws, df: pd.DataFrame):
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([_nan_to_none(x) for x in row])


def _apply_formats_by_header(ws, header_row=1):
    headers = [c.value for c in ws[header_row]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    money_headers = {
        "Neto Gravado", "IVA Liquidado", "IVA Débito", "IVA Crédito",
        "Conceptos NG/EX", "Perc./Ret.", "Total"
    }
    aliq_headers = {"Alíc."}
    cuit_headers = {"CUIT"}

    for r in range(header_row + 1, ws.max_row + 1):
        for h in money_headers:
            if h in idx:
                ws.cell(row=r, column=idx[h]).number_format = FMT_NUM2
        for h in aliq_headers:
            if h in idx:
                ws.cell(row=r, column=idx[h]).number_format = FMT_NUM3
        for h in cuit_headers:
            if h in idx:
                ws.cell(row=r, column=idx[h]).number_format = FMT_CUIT


def build_ventas_rows(liqs: List[Liquidacion]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []

    for l in liqs:
        # Línea principal
        rows.append({
            "Fecha dd/mm/aaaa": l.fecha,
            "Cpbte": l.tipo_cbte,
            "Tipo": l.letra,
            "Suc.": l.pv,
            "Número": l.numero,
            "Razón Social o Denominación Cliente ": (l.comprador.razon_social or "").strip(),
            "Tipo Doc.": 80,
            "CUIT": (l.comprador.cuit or "").replace("-", ""),
            "Domicilio": (l.comprador.domicilio or "").strip(),
            "C.P.": "",
            "Pcia": "",
            "Cond Fisc": l.comprador.cond_fisc,
            "Cód. Neto": l.cod_neto_venta,
            "Neto Gravado": float(l.neto or 0),
            "Alíc.": float(l.alic_iva or 0),
            "IVA Liquidado": float(l.iva or 0),
            "IVA Débito": float(l.iva or 0),
            "Cód. NG/EX": "",
            "Conceptos NG/EX": None,
            "Cód. P/R": "",
            "Perc./Ret.": None,
            "Pcia P/R": "",
            "Total": float(l.total or 0),
        })

        # Retenciones (IMPORTES)
        def add_ret(code: str, amount: float):
            amt = float(amount or 0)
            if amt == 0:
                return
            rows.append({
                "Fecha dd/mm/aaaa": l.fecha,
                "Cpbte": "RV",
                "Tipo": l.letra,
                "Suc.": l.pv,
                "Número": l.numero,
                "Razón Social o Denominación Cliente ": (l.comprador.razon_social or "").strip(),
                "Tipo Doc.": 80,
                "CUIT": (l.comprador.cuit or "").replace("-", ""),
                "Domicilio": (l.comprador.domicilio or "").strip(),
                "C.P.": "",
                "Pcia": "",
                "Cond Fisc": l.comprador.cond_fisc,
                "Cód. Neto": "",
                "Neto Gravado": None,
                "Alíc.": None,
                "IVA Liquidado": None,
                "IVA Débito": None,
                "Cód. NG/EX": "",
                "Conceptos NG/EX": None,
                "Cód. P/R": code,
                "Perc./Ret.": amt,
                "Pcia P/R": "",
                "Total": amt,
            })

        add_ret("RA07", l.ret_iva)
        add_ret("RA05", l.ret_gan)

    df = pd.DataFrame(rows, columns=VENTAS_COLUMNS)

    # columnas numéricas como número (no texto)
    for col in ["Neto Gravado", "Alíc.", "IVA Liquidado", "IVA Débito", "Conceptos NG/EX", "Perc./Ret.", "Total"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def build_gastos_rows(liqs: List[Liquidacion]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []

    for l in liqs:
        exento_total = 0.0
        by_alic: Dict[float, List[float]] = {}  # alic -> [neto, iva]

        for d in l.deducciones:
            if (d.alic or 0) == 0:
                exento_total += (d.total if d.total else d.neto)
            else:
                by_alic.setdefault(d.alic, [0.0, 0.0])
                by_alic[d.alic][0] += float(d.neto or 0)
                by_alic[d.alic][1] += float(d.iva or 0)

        alics_sorted = sorted(by_alic.keys())

        def add_line(neto: float, alic: Optional[float], iva: float, exento_here: float):
            mov = 202 if (alic is not None and abs(alic - 21.0) < 0.001) else 203
            total = (neto or 0) + (iva or 0) + (exento_here or 0)

            rows.append({
                "Fecha Emisión ": l.fecha,
                "Fecha Recepción": l.fecha,
                "Cpbte": "ND",
                "Suc.": l.pv,
                "Número": l.numero,
                "Razón Social/Denominación Proveedor": (l.comprador.razon_social or "").strip(),
                "Tipo Doc.": 80,
                "CUIT": (l.comprador.cuit or "").replace("-", ""),
                "Domicilio": (l.comprador.domicilio or "").strip(),
                "C.P.": "",
                "Pcia": "",
                "Cond Fisc": l.comprador.cond_fisc,
                "Cód. Neto": mov,
                "Neto Gravado": neto if neto != 0 else 0.0,
                "Alíc.": alic,
                "IVA Liquidado": iva if iva != 0 else 0.0,
                "IVA Crédito": iva if iva != 0 else 0.0,
                "Cód. NG/EX": 203 if exento_here else "",
                "Conceptos NG/EX": exento_here if exento_here else None,
                "Cód. P/R": "",
                "Perc./Ret.": None,
                "Pcia P/R": "",
                "Total": total,
            })

        if alics_sorted:
            for idx, alic in enumerate(alics_sorted):
                neto, iva = by_alic[alic]
                exento_here = exento_total if idx == 0 else 0.0
                add_line(float(neto), float(alic), float(iva), float(exento_here))
        else:
            add_line(0.0, None, 0.0, float(exento_total or 0.0))

        # Percepción IVA (P007)
        perc_iva = float(getattr(l, "perc_iva", 0.0) or 0.0)
        if perc_iva != 0.0:
            rows.append({
                "Fecha Emisión ": l.fecha,
                "Fecha Recepción": l.fecha,
                "Cpbte": "ND",
                "Suc.": l.pv,
                "Número": l.numero,
                "Razón Social/Denominación Proveedor": (l.comprador.razon_social or "").strip(),
                "Tipo Doc.": 80,
                "CUIT": (l.comprador.cuit or "").replace("-", ""),
                "Domicilio": (l.comprador.domicilio or "").strip(),
                "C.P.": "",
                "Pcia": "",
                "Cond Fisc": l.comprador.cond_fisc,
                "Cód. Neto": "",
                "Neto Gravado": None,
                "Alíc.": None,
                "IVA Liquidado": None,
                "IVA Crédito": None,
                "Cód. NG/EX": "",
                "Conceptos NG/EX": None,
                "Cód. P/R": "P007",
                "Perc./Ret.": perc_iva,
                "Pcia P/R": "",
                "Total": perc_iva,
            })

    df = pd.DataFrame(rows, columns=COMPRAS_COLUMNS)

    for col in ["Neto Gravado", "Alíc.", "IVA Liquidado", "IVA Crédito", "Conceptos NG/EX", "Perc./Ret.", "Total"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def build_excel_ventas(liqs: List[Liquidacion]) -> BytesIO:
    df = build_ventas_rows(liqs)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    _append_df(ws, df)

    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    _apply_formats_by_header(ws, 1)

    _set_col_widths(ws, [14,8,6,7,12,40,10,14,24,8,8,10,10,14,8,14,14,10,14,10,14,10,14])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_excel_gastos(liqs: List[Liquidacion]) -> BytesIO:
    df = build_gastos_rows(liqs)

    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    _append_df(ws, df)

    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    _apply_formats_by_header(ws, 1)

    _set_col_widths(ws, [14,14,8,7,12,40,10,14,24,8,8,10,10,14,8,14,14,10,14,10,14,10,14])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_excel_cpns(liqs: List[Liquidacion]) -> BytesIO:
    """
    UNA sola hoja.
    Repite filas si hay múltiples Mercadería Entregada.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "CPNs"

    headers = [
        "FECHA",
        "COE",
        "COMPROBANTE",
        "ACOPIO/COMPRADOR",
        "CUIT COMPRADOR",
        "TIPO DE GRANO",
        "CAMPAÑA",
        "KILOS",
        "PRECIO",
        "NETO",
        "ALIC IVA",
        "IVA",
        "TOTAL",
        "RET IVA",
        "RET GAN",
        "ME - Nro comprobante",
        "ME - Grado",
        "ME - Factor",
        "ME - Contenido proteico",
        "ME - Procedencia",
        "ME - Peso (kg)",
    ]
    ws.append(headers)

    bold = Font(bold=True)
    for c in ws[1]:
        c.font = bold
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for l in liqs:
        comp = f"{l.pv}-{l.numero}" if (l.pv and l.numero) else ""

        me_items = getattr(l, "me_items", None) or []
        if not me_items and (l.me_nro_comprobante or ""):
            me_items = [{
                "nro": l.me_nro_comprobante,
                "grado": l.me_grado,
                "factor": l.me_factor,
                "prot": l.me_contenido_proteico,
                "peso": l.me_peso_kg,
                "proced": l.me_procedencia,
            }]
        if not me_items:
            me_items = [{}]

        for it in me_items:
            ws.append([
                l.fecha or "",
                l.coe or "",
                comp,
                (l.acopio.razon_social or "").strip(),
                (l.comprador.cuit or "").replace("-", ""),
                l.grano or "",
                l.campaña or "",
                float(l.kilos or 0),
                float(l.precio or 0),
                float(l.neto or 0),
                float(l.alic_iva or 0),
                float(l.iva or 0),
                float(l.total or 0),
                float(l.ret_iva or 0),
                float(l.ret_gan or 0),
                it.get("nro", "") or "",
                it.get("grado", "") or "",
                _nan_to_none(it.get("factor", None)),
                _nan_to_none(it.get("prot", None)),
                it.get("proced", "") or "",
                _nan_to_none(it.get("peso", None)),
            ])

    idx = {h: i + 1 for i, h in enumerate(headers)}
    for r in range(2, ws.max_row + 1):
        ws.cell(r, idx["KILOS"]).number_format = FMT_NUM2
        ws.cell(r, idx["PRECIO"]).number_format = FMT_MONEY
        ws.cell(r, idx["NETO"]).number_format = FMT_NUM2
        ws.cell(r, idx["ALIC IVA"]).number_format = FMT_NUM3
        ws.cell(r, idx["IVA"]).number_format = FMT_NUM2
        ws.cell(r, idx["TOTAL"]).number_format = FMT_NUM2
        ws.cell(r, idx["RET IVA"]).number_format = FMT_NUM2
        ws.cell(r, idx["RET GAN"]).number_format = FMT_NUM2
        ws.cell(r, idx["ME - Factor"]).number_format = FMT_NUM2
        ws.cell(r, idx["ME - Contenido proteico"]).number_format = FMT_NUM2
        ws.cell(r, idx["ME - Peso (kg)"]).number_format = FMT_NUM2

    _set_col_widths(ws, [12,14,16,40,14,18,12,14,14,14,10,14,14,12,12,18,12,12,20,22,14])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
